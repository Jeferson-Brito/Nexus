"""
API REST para o Sistema de Controle de Ponto do Nexus.

Endpoints:
  POST /api/ponto/buscar-colaborador/   → Busca colaborador pelos 6 primeiros dígitos do CPF
  POST /api/ponto/registrar/            → Registra ponto (tablet) com foto base64
  GET  /api/ponto/registros/            → Lista registros com filtros
  POST /api/ponto/ajustar/<pk>/         → Ajuste manual de registro (admin)
  GET  /api/ponto/dashboard/            → KPIs do dia para o painel admin
  GET  /api/ponto/banco-horas/<pk>/     → Banco de horas de um colaborador
  GET  /api/ponto/relatorio/            → Relatório mensal
  GET  /api/ponto/exportar-excel/       → Exportação Excel
"""
import base64
import json
import io
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from ..models import (
    Colaborador, RegistroPonto, BancoHoras,
    ConfiguracaoPonto, Department, User, Holiday
)


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _is_admin_or_gestor(user):
    return user.role in ('administrador', 'gestor')


def _calcular_horas_dia(registros):
    """
    Dado um queryset de RegistroPonto do mesmo dia (ordenados por hora),
    retorna os minutos trabalhados no dia.
    """
    mapa = {r.tipo: r.hora for r in registros}

    entrada = mapa.get('entrada')
    saida = mapa.get('saida')
    saida_almoco = mapa.get('saida_almoco')
    retorno_almoco = mapa.get('retorno_almoco')

    if not entrada or not saida:
        return 0

    def to_minutes(t):
        return t.hour * 60 + t.minute

    total = to_minutes(saida) - to_minutes(entrada)

    # Desconta intervalo de almoço se batido
    if saida_almoco and retorno_almoco:
        almoco = to_minutes(retorno_almoco) - to_minutes(saida_almoco)
        total -= max(almoco, 0)

    return max(total, 0)


def _atualizar_banco_horas(colaborador, data_ref=None):
    """Recalcula e salva o banco de horas de um colaborador."""
    if data_ref is None:
        data_ref = date.today()

    # Config de ponto do departamento
    config = ConfiguracaoPonto.objects.filter(
        department=colaborador.department, ativo=True
    ).first()
    carga_diaria = config.carga_horaria_diaria if config else 480  # 8h

    # Todos os registros do mês corrente
    registros = RegistroPonto.objects.filter(
        colaborador=colaborador,
        data__year=data_ref.year,
        data__month=data_ref.month,
    ).order_by('data', 'hora')

    dias = {}
    for r in registros:
        dias.setdefault(r.data, []).append(r)

    # Buscar feriados do mês para evitar múltiplas queries se necessário
    # Mas como _atualizar_banco_horas costuma ser para um colaborador/mês, podemos otimizar
    feriados = Holiday.objects.filter(
        Q(date__year=data_ref.year, date__month=data_ref.month) | 
        Q(repeats_annually=True, date__month=data_ref.month)
    )
    datas_feriados = {f.date.replace(year=data_ref.year) if f.repeats_annually else f.date for f in feriados}

    total_trabalhado = 0
    total_esperado = 0

    # Iterar pelos dias que tiveram registro
    for data_dia, regs_dia in dias.items():
        trabalhado_dia = _calcular_horas_dia(regs_dia)
        total_trabalhado += trabalhado_dia
        
        # Se for feriado, carga esperada é 0 (trabalho no feriado conta como extra)
        if data_dia in datas_feriados:
            total_esperado += 0
        else:
            total_esperado += carga_diaria

    saldo = total_trabalhado - total_esperado

    banco, _ = BancoHoras.objects.get_or_create(colaborador=colaborador)
    banco.saldo_minutos = saldo
    banco.save()
    return banco


# ─────────────────────────────────────────────────────────────
# Endpoint: Buscar colaborador por CPF (6 primeiros dígitos)
# ─────────────────────────────────────────────────────────────

@csrf_exempt
@require_http_methods(['POST'])
def buscar_colaborador(request):
    """Recebe os 6 primeiros dígitos do CPF e retorna dados do colaborador."""
    try:
        data = json.loads(request.body)
        cpf_iniciais = data.get('cpf_iniciais', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'erro': 'JSON inválido'}, status=400)

    if len(cpf_iniciais) != 6 or not cpf_iniciais.isdigit():
        return JsonResponse({'erro': 'Informe exatamente 6 dígitos do CPF'}, status=400)

    # CPF armazenado pode ter formatação: 000.000.000-00 → remover pontos e traço
    colaboradores = Colaborador.objects.filter(status='ativo').select_related('department')
    encontrado = None
    for col in colaboradores:
        cpf_limpo = col.cpf.replace('.', '').replace('-', '')
        if cpf_limpo.startswith(cpf_iniciais):
            encontrado = col
            break

    if not encontrado:
        return JsonResponse({'erro': 'Colaborador não encontrado ou inativo'}, status=404)

    # Foto
    foto_url = None
    if encontrado.foto:
        foto_url = request.build_absolute_uri(encontrado.foto.url)

    # Tipos já registrados hoje (para desabilitar botões no kiosk)
    from datetime import date as _date
    hoje = _date.today()
    tipos_hoje = list(
        RegistroPonto.objects.filter(
            colaborador=encontrado, data=hoje
        ).values_list('tipo', flat=True)
    )

    return JsonResponse({
        'id': str(encontrado.id),
        'nome': encontrado.nome_completo,
        'cargo': encontrado.cargo_atual,
        'departamento': encontrado.department.name if encontrado.department else '',
        'foto_url': foto_url,
        'tipos_hoje': tipos_hoje,
        'exigir_foto': encontrado.ponto_web_foto,
    })


# ─────────────────────────────────────────────────────────────
# Endpoint: Registrar ponto
# ─────────────────────────────────────────────────────────────

@csrf_exempt
@require_http_methods(['POST'])
def registrar_ponto(request):
    """
    Registra o ponto do colaborador. Aceita foto em base64.
    Payload: { colaborador_id, tipo, foto_base64 }
    """
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'erro': 'JSON inválido'}, status=400)

    colaborador_id = data.get('colaborador_id')
    tipo = data.get('tipo')
    foto_base64 = data.get('foto_base64', '')
    origem = data.get('origem', 'tablet')

    if not colaborador_id or not tipo:
        return JsonResponse({'erro': 'colaborador_id e tipo são obrigatórios'}, status=400)

    TIPOS_VALIDOS = ['entrada', 'saida_almoco', 'retorno_almoco', 'saida']
    if tipo not in TIPOS_VALIDOS:
        return JsonResponse({'erro': f'Tipo inválido. Use: {TIPOS_VALIDOS}'}, status=400)

    try:
        colaborador = Colaborador.objects.get(id=int(colaborador_id), status='ativo')
    except Colaborador.DoesNotExist:
        return JsonResponse({'erro': f'Colaborador não encontrado para id "{colaborador_id}".'}, status=404)
    except ValueError:
        return JsonResponse({'erro': f'ValueError: id "{colaborador_id}" não é inteiro.'}, status=404)
    except TypeError:
        return JsonResponse({'erro': f'TypeError: id "{colaborador_id}" é inválido.'}, status=404)
    except Exception as e:
        return JsonResponse({'erro': f'Unknown Error: {str(e)}'}, status=500)

    agora = timezone.localtime(timezone.now())
    data_hoje = agora.date()
    hora_agora = agora.time()

    # Verificar registro duplicado (mesmo tipo no mesmo dia)
    if RegistroPonto.objects.filter(
        colaborador=colaborador, data=data_hoje, tipo=tipo
    ).exists():
        return JsonResponse({
            'erro': f'Registro de "{tipo}" já realizado hoje'
        }, status=409)

    # Salvar foto
    foto_file = None
    if foto_base64:
        try:
            if ',' in foto_base64:
                foto_base64 = foto_base64.split(',')[1]
            foto_bytes = base64.b64decode(foto_base64)
            nome_arquivo = f'ponto_{colaborador.id}_{data_hoje}_{tipo}.jpg'
            foto_file = ContentFile(foto_bytes, name=nome_arquivo)
        except Exception:
            pass  # Foto inválida → registra sem foto

    # Criar registro
    registrado_por = None
    if request.user.is_authenticated:
        registrado_por = request.user

    registro = RegistroPonto.objects.create(
        colaborador=colaborador,
        tipo=tipo,
        data=data_hoje,
        hora=hora_agora,
        foto=foto_file,
        origem=origem,
        registrado_por=registrado_por,
    )

    # Atualizar banco de horas
    try:
        _atualizar_banco_horas(colaborador, data_hoje)
    except Exception:
        pass

    return JsonResponse({
        'sucesso': True,
        'id': registro.id,
        'tipo': registro.get_tipo_display(),
        'data': data_hoje.isoformat(),
        'hora': hora_agora.strftime('%H:%M'),
        'mensagem': f'Ponto de {registro.get_tipo_display()} registrado com sucesso!',
    })


# ─────────────────────────────────────────────────────────────
# Endpoint: Listar registros
# ─────────────────────────────────────────────────────────────

@login_required
def listar_registros(request):
    """Lista registros com filtros opcionais."""
    user = request.user

    qs = RegistroPonto.objects.select_related(
        'colaborador', 'colaborador__department', 'registrado_por'
    )

    # Filtros de acesso
    if not _is_admin_or_gestor(user):
        # Colaborador só vê os próprios registros
        try:
            colaborador = user.colaborador_perfil
            qs = qs.filter(colaborador=colaborador)
        except AttributeError:
            return JsonResponse({'registros': []})
    else:
        # Admin/Gestor filtra por departamento se informado
        dept_id = request.GET.get('department_id')
        if dept_id:
            qs = qs.filter(colaborador__department_id=dept_id)
        elif user.role == 'gestor' and user.department:
            qs = qs.filter(colaborador__department=user.department)

    # Filtros de data
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    colaborador_id = request.GET.get('colaborador_id')

    if data_inicio:
        qs = qs.filter(data__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data__lte=data_fim)
    if colaborador_id:
        qs = qs.filter(colaborador_id=colaborador_id)

    # Limite
    qs = qs.order_by('-data', '-hora')[:200]

    registros = []
    for r in qs:
        foto_url = None
        if r.foto:
            try:
                foto_url = request.build_absolute_uri(r.foto.url)
            except Exception:
                pass

        registros.append({
            'id': r.id,
            'colaborador_id': r.colaborador_id,
            'colaborador_nome': r.colaborador.nome_completo,
            'departamento': r.colaborador.department.name if r.colaborador.department else '',
            'tipo': r.tipo,
            'tipo_display': r.get_tipo_display(),
            'data': r.data.isoformat(),
            'hora': r.hora.strftime('%H:%M'),
            'origem': r.get_origem_display(),
            'foto_url': foto_url,
            'observacao': r.observacao,
        })

    return JsonResponse({'registros': registros})


# ─────────────────────────────────────────────────────────────
# Endpoint: Ajuste manual (admin)
# ─────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['POST'])
def ajustar_registro(request, pk=None):
    """
    Ajusta ou cria manualmente um registro de ponto.
    Payload: { colaborador_id, tipo, data, hora, observacao }
    """
    if not _is_admin_or_gestor(request.user):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'erro': 'JSON inválido'}, status=400)

    if pk:
        # Editar registro existente
        try:
            registro = RegistroPonto.objects.get(id=pk)
        except RegistroPonto.DoesNotExist:
            return JsonResponse({'erro': 'Registro não encontrado'}, status=404)
        registro.hora = data.get('hora', registro.hora)
        registro.observacao = data.get('observacao', registro.observacao)
        registro.origem = 'admin'
        registro.registrado_por = request.user
        registro.save()
    else:
        # Criar novo registro manual
        colaborador_id = data.get('colaborador_id')
        tipo = data.get('tipo')
        data_str = data.get('data', date.today().isoformat())
        hora_str = data.get('hora', '00:00')
        obs = data.get('observacao', 'Lançamento manual')

        try:
            colaborador = Colaborador.objects.get(id=colaborador_id)
        except Colaborador.DoesNotExist:
            return JsonResponse({'erro': 'Colaborador não encontrado'}, status=404)

        data_reg = datetime.strptime(data_str, '%Y-%m-%d').date()
        hora_reg = datetime.strptime(hora_str, '%H:%M').time()

        registro = RegistroPonto.objects.create(
            colaborador=colaborador,
            tipo=tipo,
            data=data_reg,
            hora=hora_reg,
            origem='admin',
            registrado_por=request.user,
            observacao=obs,
        )

        # Atualizar banco de horas
        try:
            _atualizar_banco_horas(colaborador, data_reg)
        except Exception:
            pass

    return JsonResponse({'sucesso': True, 'id': registro.id})


@login_required
@require_http_methods(['DELETE'])
def deletar_registro(request, pk):
    """Remove um registro de ponto (admin only)."""
    if not _is_admin_or_gestor(request.user):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)

    try:
        registro = RegistroPonto.objects.get(id=pk)
        colaborador = registro.colaborador
        data_reg = registro.data
        registro.delete()
        _atualizar_banco_horas(colaborador, data_reg)
        return JsonResponse({'sucesso': True})
    except RegistroPonto.DoesNotExist:
        return JsonResponse({'erro': 'Registro não encontrado'}, status=404)


# ─────────────────────────────────────────────────────────────
# Endpoint: Dashboard do dia
# ─────────────────────────────────────────────────────────────

@login_required
def dashboard_ponto(request):
    """KPIs do dia para o painel administrativo."""
    if not _is_admin_or_gestor(request.user):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)

    hoje = timezone.localtime(timezone.now()).date()

    # Filtro por departamento
    dept_id = request.GET.get('department_id')
    if dept_id:
        colaboradores_qs = Colaborador.objects.filter(
            status='ativo', department_id=dept_id
        )
    elif request.user.role == 'gestor' and request.user.department:
        colaboradores_qs = Colaborador.objects.filter(
            status='ativo', department=request.user.department
        )
    else:
        colaboradores_qs = Colaborador.objects.filter(status='ativo')

    total_colaboradores = colaboradores_qs.count()

    # Quem bateu entrada hoje
    entradas_hoje = set(
        RegistroPonto.objects.filter(
            data=hoje, tipo='entrada',
            colaborador__in=colaboradores_qs
        ).values_list('colaborador_id', flat=True)
    )

    # Quem bateu saída final hoje
    saidas_hoje = set(
        RegistroPonto.objects.filter(
            data=hoje, tipo='saida',
            colaborador__in=colaboradores_qs
        ).values_list('colaborador_id', flat=True)
    )

    presentes = len(entradas_hoje - saidas_hoje)
    ausentes = total_colaboradores - len(entradas_hoje)
    finalizaram = len(saidas_hoje)

    # Últimos 10 registros do dia
    ultimos = RegistroPonto.objects.filter(
        data=hoje, colaborador__in=colaboradores_qs
    ).select_related('colaborador').order_by('-hora')[:10]

    feed = []
    for r in ultimos:
        foto_url = None
        if r.foto:
            try:
                foto_url = request.build_absolute_uri(r.foto.url)
            except Exception:
                pass
        feed.append({
            'colaborador': r.colaborador.nome_completo,
            'tipo': r.get_tipo_display(),
            'hora': r.hora.strftime('%H:%M'),
            'foto_url': foto_url,
        })

    return JsonResponse({
        'data': hoje.isoformat(),
        'total_colaboradores': total_colaboradores,
        'presentes': presentes,
        'ausentes': ausentes,
        'finalizaram': finalizaram,
        'feed': feed,
    })


# ─────────────────────────────────────────────────────────────
# Endpoint: Banco de horas
# ─────────────────────────────────────────────────────────────

@login_required
def banco_horas(request, pk):
    """Retorna o banco de horas de um colaborador."""
    try:
        colaborador = Colaborador.objects.get(id=pk)
    except Colaborador.DoesNotExist:
        return JsonResponse({'erro': 'Colaborador não encontrado'}, status=404)

    # Permissão: admin/gestor ou o próprio colaborador
    if not _is_admin_or_gestor(request.user):
        try:
            if request.user.colaborador_perfil.id != pk:
                return JsonResponse({'erro': 'Sem permissão'}, status=403)
        except AttributeError:
            return JsonResponse({'erro': 'Sem permissão'}, status=403)

    banco, _ = BancoHoras.objects.get_or_create(colaborador=colaborador)

    return JsonResponse({
        'colaborador_id': pk,
        'colaborador_nome': colaborador.nome_completo,
        'saldo_minutos': banco.saldo_minutos,
        'saldo_formatado': banco.saldo_formatado,
        'updated_at': banco.updated_at.isoformat(),
    })


# ─────────────────────────────────────────────────────────────
# Endpoint: Relatório mensal
# ─────────────────────────────────────────────────────────────

@login_required
def relatorio_mensal(request):
    """Retorna relatório consolidado por colaborador para um mês/ano."""
    if not _is_admin_or_gestor(request.user):
        return JsonResponse({'erro': 'Sem permissão'}, status=403)

    hoje = date.today()
    mes = int(request.GET.get('mes', hoje.month))
    ano = int(request.GET.get('ano', hoje.year))
    dept_id = request.GET.get('department_id')

    colaboradores_qs = Colaborador.objects.filter(status='ativo').select_related('department')
    if dept_id:
        colaboradores_qs = colaboradores_qs.filter(department_id=dept_id)
    elif request.user.role == 'gestor' and request.user.department:
        colaboradores_qs = colaboradores_qs.filter(department=request.user.department)

    relatorio = []

    for col in colaboradores_qs:
        config = ConfiguracaoPonto.objects.filter(
            department=col.department, ativo=True
        ).first()
        carga_diaria = config.carga_horaria_diaria if config else 480

        registros = RegistroPonto.objects.filter(
            colaborador=col, data__year=ano, data__month=mes
        ).order_by('data', 'hora')

        dias = {}
        for r in registros:
            dias.setdefault(r.data, []).append(r)

        total_min = sum(_calcular_horas_dia(v) for v in dias.values())
        dias_presentes = len(dias)
        esperado_min = carga_diaria * dias_presentes
        saldo_min = total_min - esperado_min

        def fmt(m):
            h = abs(m) // 60
            mi = abs(m) % 60
            s = '+' if m >= 0 else '-'
            return f'{s}{h:02d}:{mi:02d}'

        relatorio.append({
            'colaborador_id': col.id,
            'colaborador_nome': col.nome_completo,
            'departamento': col.department.name if col.department else '',
            'dias_presentes': dias_presentes,
            'total_trabalhado': fmt(total_min),
            'total_esperado': fmt(esperado_min),
            'saldo': fmt(saldo_min),
            'saldo_minutos': saldo_min,
        })

    return JsonResponse({'mes': mes, 'ano': ano, 'relatorio': relatorio})


# ─────────────────────────────────────────────────────────────
# Endpoint: Exportar Excel
# ─────────────────────────────────────────────────────────────

@login_required
def exportar_excel(request):
    """Gera e retorna um arquivo Excel com o relatório mensal de ponto."""
    if not _is_admin_or_gestor(request.user):
        return HttpResponse('Sem permissão', status=403)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    hoje = date.today()
    mes = int(request.GET.get('mes', hoje.month))
    ano = int(request.GET.get('ano', hoje.year))
    dept_id = request.GET.get('department_id')

    colaboradores_qs = Colaborador.objects.filter(status='ativo').select_related('department')
    if dept_id:
        colaboradores_qs = colaboradores_qs.filter(department_id=dept_id)
    elif request.user.role == 'gestor' and request.user.department:
        colaboradores_qs = colaboradores_qs.filter(department=request.user.department)

    wb = Workbook()
    ws = wb.active
    ws.title = f'Ponto {mes:02d}/{ano}'

    # Cabeçalho
    headers = ['Colaborador', 'Departamento', 'Dias Presentes', 'Total Trabalhado', 'Total Esperado', 'Saldo']
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Dados
    row = 2
    for col in colaboradores_qs:
        config = ConfiguracaoPonto.objects.filter(
            department=col.department, ativo=True
        ).first()
        carga_diaria = config.carga_horaria_diaria if config else 480

        registros = RegistroPonto.objects.filter(
            colaborador=col, data__year=ano, data__month=mes
        ).order_by('data', 'hora')

        dias = {}
        for r in registros:
            dias.setdefault(r.data, []).append(r)

        total_min = sum(_calcular_horas_dia(v) for v in dias.values())
        dias_presentes = len(dias)
        esperado_min = carga_diaria * dias_presentes
        saldo_min = total_min - esperado_min

        def fmt(m):
            h = abs(m) // 60
            mi = abs(m) % 60
            s = '+' if m >= 0 else '-'
            return f'{s}{h:02d}:{mi:02d}'

        ws.cell(row=row, column=1, value=col.nome_completo)
        ws.cell(row=row, column=2, value=col.department.name if col.department else '')
        ws.cell(row=row, column=3, value=dias_presentes)
        ws.cell(row=row, column=4, value=fmt(total_min))
        ws.cell(row=row, column=5, value=fmt(esperado_min))
        cell_saldo = ws.cell(row=row, column=6, value=fmt(saldo_min))
        if saldo_min < 0:
            cell_saldo.font = Font(color='C0392B')
        else:
            cell_saldo.font = Font(color='27AE60')
        row += 1

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    # Resposta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ponto_{mes:02d}_{ano}.xlsx"'
    return response

import csv
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from ..models import User

@login_required
def export_users_csv(request):
    """Exportar usuários para CSV - apenas gestores"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para exportar usuários.')
        return redirect('dashboard')
    
    # Filtro por departamento
    if request.user.is_administrador():
        users = User.objects.all().order_by('username')
    else:
        users = User.objects.filter(department=request.user.department).order_by('username')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Username', 'E-mail', 'Nome', 'Sobrenome', 'Perfil', 'Ativo', 'Último Login', 'Data de Criação'])
    
    for user in users:
        writer.writerow([
            user.username,
            user.email,
            user.first_name or '',
            user.last_name or '',
            user.get_role_display(),
            'Sim' if user.ativo else 'Não',
            user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else '',
            user.date_joined.strftime('%d/%m/%Y %H:%M') if user.date_joined else ''
        ])
    
    return response

@login_required
def export_users_xlsx(request):
    """Exportar usuários para XLSX - apenas gestores e administradores"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para exportar usuários.')
        return redirect('dashboard')
    
    # Filtro por departamento
    if request.user.is_administrador():
        users = User.objects.all().order_by('username')
    else:
        users = User.objects.filter(department=request.user.department).order_by('username')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuários"
    
    headers = ['Username', 'E-mail', 'Nome', 'Sobrenome', 'Perfil', 'Ativo', 'Último Login', 'Data de Criação']
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, user in enumerate(users, 2):
        ws.cell(row=row_num, column=1, value=user.username)
        ws.cell(row=row_num, column=2, value=user.email)
        ws.cell(row=row_num, column=3, value=user.first_name or '')
        ws.cell(row=row_num, column=4, value=user.last_name or '')
        ws.cell(row=row_num, column=5, value=user.get_role_display())
        ws.cell(row=row_num, column=6, value='Sim' if user.ativo else 'Não')
        ws.cell(row=row_num, column=7, value=user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else '')
        ws.cell(row=row_num, column=8, value=user.date_joined.strftime('%d/%m/%Y %H:%M') if user.date_joined else '')
    
    column_widths = [20, 30, 20, 20, 15, 10, 20, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

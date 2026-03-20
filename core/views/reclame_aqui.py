from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg
from django.http import JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta
import re
from ..models import Complaint, Store, User, Department, Activity, AuditLog
from ..forms import ComplaintForm

@login_required
def complaint_list(request):
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            complaints = Complaint.objects.filter(department_id=selected_dept_id).select_related('analista', 'department')
        else:
            complaints = Complaint.objects.all().select_related('analista', 'department')
    else:
        complaints = Complaint.objects.filter(department=request.user.department).select_related('analista')
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    loja_filter = request.GET.get('loja', '')
    
    if search:
        # Remover formatação do CPF se houver (pontos e traços)
        search_clean = re.sub(r'[^\d\w\s@.-]', '', search)  # Remove apenas pontos e traços, mantém números e letras
        
        # Se parecer um CPF (11 dígitos), buscar apenas números
        numbers_only = re.sub(r'\D', '', search_clean)
        if len(numbers_only) == 11:
            # Buscar CPF sem formatação
            complaints = complaints.filter(
                Q(id_ra__icontains=search) |
                Q(cpf_cliente__icontains=numbers_only) |
                Q(nome_cliente__icontains=search) |
                Q(email_cliente__icontains=search)
            )
        else:
            # Busca normal
            complaints = complaints.filter(
                Q(id_ra__icontains=search) |
                Q(cpf_cliente__icontains=search_clean) |
                Q(nome_cliente__icontains=search) |
                Q(email_cliente__icontains=search)
            )
    
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    if loja_filter:
        complaints = complaints.filter(loja_cod=loja_filter)
    
    # Filtro de reclamações urgentes (pendentes há mais de 3 dias)
    urgentes = request.GET.get('urgentes', '')
    if urgentes == 'true':
        urgent_date = timezone.now().date() - timedelta(days=3)
        complaints = complaints.filter(
            status='pendente',
            data_reclamacao__lte=urgent_date
        )
    
    # Filtros avançados
    tipo_filter = request.GET.get('tipo', '')
    analista_filter = request.GET.get('analista', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    sem_analista = request.GET.get('sem_analista', '')
    
    if tipo_filter:
        complaints = complaints.filter(tipo_reclamacao=tipo_filter)
    
    if analista_filter:
        if analista_filter == 'sem_analista':
            complaints = complaints.filter(analista__isnull=True)
        else:
            complaints = complaints.filter(analista_id=analista_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            complaints = complaints.filter(data_reclamacao__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            complaints = complaints.filter(data_reclamacao__lte=date_to_obj)
        except:
            pass
    
    if sem_analista == 'true':
        complaints = complaints.filter(analista__isnull=True)
    
    complaints = complaints.order_by('-created_at')
    
    # Contador de reclamações por responsável - analistas e gestores do depto
    if request.user.is_administrador():
        if selected_dept_id:
            base_analysts_stats = Complaint.objects.filter(department_id=selected_dept_id)
            analistas_list = User.objects.filter(role__in=['analista', 'gestor'], ativo=True, department_id=selected_dept_id).order_by('first_name')
        else:
            base_analysts_stats = Complaint.objects.all()
            analistas_list = User.objects.filter(role__in=['analista', 'gestor'], ativo=True).order_by('first_name')
    else:
        base_analysts_stats = Complaint.objects.filter(department=request.user.department)
        analistas_list = User.objects.filter(role__in=['analista', 'gestor'], ativo=True, department=request.user.department).order_by('first_name')

    complaints_by_analyst = base_analysts_stats.filter(
        analista__isnull=False
    ).values('analista__username', 'analista__first_name', 'analista__last_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    paginator = Paginator(complaints, 25)
    page = request.GET.get('page')
    complaints = paginator.get_page(page)
    
    context = {
        'complaints': complaints,
        'complaints_by_analyst': complaints_by_analyst,
        'analistas_list': analistas_list,
    }
    
    return render(request, 'core/complaint_list.html', context)

@login_required
def store_list(request):
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            queryset = Complaint.objects.filter(department_id=selected_dept_id)
        else:
            queryset = Complaint.objects.all()
    else:
        queryset = Complaint.objects.filter(department=request.user.department)

    """Lista todas as lojas com reclamações, com filtros e ordenação"""
    # Obter todas as lojas com contagem de reclamações
    stores = queryset.values('loja_cod').annotate(
        count=Count('id'),
        pendentes=Count('id', filter=Q(status='pendente')),
        em_andamento=Count('id', filter=Q(status='em_andamento')),
        resolvidas=Count('id', filter=Q(status='resolvido')),
        aguardando=Count('id', filter=Q(status='aguardando_avaliacao'))
    )
    
    # Filtros
    search = request.GET.get('search', '')
    min_occurrences = request.GET.get('min_occurrences', '')
    
    if search:
        stores = stores.filter(loja_cod__icontains=search)
    
    if min_occurrences:
        try:
            min_count = int(min_occurrences)
            stores = stores.filter(count__gte=min_count)
        except ValueError:
            pass
    
    # Ordenação
    order_by = request.GET.get('order_by', 'count')
    order_direction = request.GET.get('order_direction', 'desc')
    
    if order_by == 'loja':
        if order_direction == 'asc':
            stores = stores.order_by('loja_cod')
        else:
            stores = stores.order_by('-loja_cod')
    elif order_by == 'count':
        if order_direction == 'asc':
            stores = stores.order_by('count')
        else:
            stores = stores.order_by('-count')
    else:
        stores = stores.order_by('-count')
    
    # Contar total antes da paginação
    total_stores = stores.count()
    
    # Paginação - aplicar limit e offset manualmente
    page = request.GET.get('page', 1)
    try:
        page = int(page)
    except (ValueError, TypeError):
        page = 1
    
    per_page = 20
    start = (page - 1) * per_page
    end = start + per_page
    
    stores_list = list(stores[start:end])
    
    # Calcular informações de paginação
    total_pages = (total_stores + per_page - 1) // per_page
    
    context = {
        'stores': stores_list,
        'search': search,
        'min_occurrences': min_occurrences,
        'order_by': order_by,
        'order_direction': order_direction,
        'total_stores': total_stores,
        'page': page,
        'total_pages': total_pages,
        'has_previous': page > 1,
        'has_next': page < total_pages,
        'previous_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < total_pages else None,
    }
    
    return render(request, 'core/store_list.html', context)

@login_required
def store_complaints(request, loja_cod):
    """Lista todas as reclamações de uma loja específica"""
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            base_queryset = Complaint.objects.filter(department_id=selected_dept_id)
        else:
            base_queryset = Complaint.objects.all()
    else:
        base_queryset = Complaint.objects.filter(department=request.user.department)

    complaints = base_queryset.filter(loja_cod=loja_cod).select_related('analista').order_by('-created_at')
    
    # Estatísticas da loja
    store_stats = base_queryset.filter(loja_cod=loja_cod).aggregate(
        total=Count('id'),
        pendentes=Count('id', filter=Q(status='pendente')),
        em_andamento=Count('id', filter=Q(status='em_andamento')),
        resolvidas=Count('id', filter=Q(status='resolvido')),
        aguardando=Count('id', filter=Q(status='aguardando_avaliacao'))
    )
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    if search:
        complaints = complaints.filter(
            Q(id_ra__icontains=search) |
            Q(cpf_cliente__icontains=search) |
            Q(nome_cliente__icontains=search) |
            Q(email_cliente__icontains=search)
        )
    
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    # Paginação
    paginator = Paginator(complaints, 25)
    page = request.GET.get('page')
    complaints_page = paginator.get_page(page)
    
    context = {
        'loja_cod': loja_cod,
        'complaints': complaints_page,
        'store_stats': store_stats,
        'search': search,
        'status_filter': status_filter,
    }
    
    return render(request, 'core/store_complaints.html', context)

@login_required
def complaint_detail(request, pk):
    # Otimização: carregar departamento e analista em uma única query
    complaint = get_object_or_404(
        Complaint.objects.select_related('department', 'analista'), 
        pk=pk
    )
    
    # Trava de segurança por departamento
    if not request.user.is_administrador():
        if complaint.department != request.user.department:
            messages.error(request, 'Você não tem permissão para acessar esta reclamação.')
            return redirect('dashboard')
            
    # Atividades já estão com select_related('usuario')
    activities = complaint.activities.select_related('usuario').order_by('-created_at')

    # Adicionar comentário interno rápido
    if request.method == 'POST' and 'comentario_interno' in request.POST:
        comentario = request.POST.get('comentario_interno', '').strip()
        if comentario:
            Activity.objects.create(
                complaint=complaint,
                usuario=request.user,
                comentario=comentario,
                tipo_interacao='comentario_interno'
            )
            messages.success(request, 'Comentário adicionado com sucesso!')
            return redirect('complaint_detail', pk=pk)
    
    # Mudança rápida de status
    if request.method == 'POST' and 'novo_status' in request.POST:
        novo_status = request.POST.get('novo_status', '').strip()
        if novo_status and novo_status != complaint.status:
            status_antigo = complaint.get_status_display()
            complaint.status = novo_status
            complaint.save()
            Activity.objects.create(
                complaint=complaint,
                usuario=request.user,
                comentario=f'Status alterado de "{status_antigo}" para "{complaint.get_status_display()}"',
                tipo_interacao='mudanca_status'
            )
            messages.success(request, f'Status alterado para "{complaint.get_status_display()}"!')
            return redirect('complaint_detail', pk=pk)
    
    # Atribuição rápida de analista
    if request.method == 'POST' and 'novo_analista' in request.POST:
        # Atribuição deve respeitar o departamento
        novo_analista_id = request.POST.get('novo_analista', '').strip()
        if novo_analista_id:
            try:
                from django.contrib.auth import authenticate # Not strictly needed here but following logic
                if request.user.is_administrador():
                    novo_analista = User.objects.get(id=novo_analista_id, role__in=['analista', 'gestor'], ativo=True, department=complaint.department)
                else:
                    novo_analista = User.objects.get(id=novo_analista_id, role__in=['analista', 'gestor'], ativo=True, department=request.user.department)
                    
                analista_antigo = complaint.analista.username if complaint.analista else "Não atribuído"
                complaint.analista = novo_analista
                complaint.save()
                Activity.objects.create(
                    complaint=complaint,
                    usuario=request.user,
                    comentario=f'Responsável alterado de "{analista_antigo}" para "{novo_analista.username}"',
                    tipo_interacao='atualizacao'
                )
                messages.success(request, f'Responsável atribuído: {novo_analista.username}!')
                return redirect('complaint_detail', pk=pk)
            except User.DoesNotExist:
                messages.error(request, 'Responsável não encontrado ou pertence a outro departamento!')
    
    # Lista de responsáveis para atribuição rápida (analistas e gestores do depto)
    # Otimização: carregar departamento para evitar query no template (se houver loop)
    if request.user.is_administrador():
        analistas_list = User.objects.filter(
            role__in=['analista', 'gestor'], 
            ativo=True, 
            department=complaint.department
        ).select_related('department').order_by('first_name')
    else:
        analistas_list = User.objects.filter(
            role__in=['analista', 'gestor'], 
            ativo=True, 
            department=request.user.department
        ).select_related('department').order_by('first_name')
    
    response = render(request, 'core/complaint_detail.html', {
        'complaint': complaint,
        'activities': activities,
        'analistas_list': analistas_list,
    })
    return response

@login_required
def complaint_create(request):
    # Trava de segurança: apenas Administradores ou o departamento 'CS Clientes' podem criar reclamações
    if not request.user.is_administrador():
        if not request.user.department or request.user.department.name != 'CS Clientes':
            messages.error(request, 'Apenas o departamento de CS Clientes e administradores podem criar reclamações.')
            # Redirecionar para a lista de reclamações ou dashboard
            return redirect('complaint_list')

    if request.method == 'POST':
        form = ComplaintForm(request.POST, user=request.user)
        if form.is_valid():
            complaint = form.save(commit=False)
            
            # Definir o departamento
            if request.user.is_administrador():
                selected_dept_id = request.session.get('selected_department_id')
                if selected_dept_id:
                    from .models import Department
                    complaint.department = Department.objects.filter(id=selected_dept_id).first()
                else:
                    # Se for admin e não tiver depto na sessão, buscar o padrão NRS
                    from .models import Department
                    complaint.department = Department.objects.filter(name='NRS Suporte').first()
            else:
                # Outros usuários usam seu departamento fixo
                complaint.department = request.user.department
            
            if request.user.is_analista():
                complaint.analista = request.user
            
            complaint.save()
            
            Activity.objects.create(
                complaint=complaint,
                usuario=request.user,
                comentario='Reclamação criada',
                tipo_interacao='criacao'
            )
            AuditLog.objects.create(
                usuario=request.user,
                action='create',
                target_type='Complaint',
                target_id=complaint.id
            )
            messages.success(request, 'Reclamação criada com sucesso!')
            return redirect('complaint_detail', pk=complaint.pk)
    else:
        form = ComplaintForm(user=request.user)
    return render(request, 'core/complaint_form.html', {'form': form})

@login_required
def complaint_edit(request, pk):
    from django.contrib.auth import authenticate
    complaint = get_object_or_404(Complaint, pk=pk)
    
    # Trava de segurança por departamento
    if not request.user.is_administrador():
        if complaint.department != request.user.department:
            messages.error(request, 'Você não tem permissão para editar esta reclamação.')
            return redirect('dashboard')
            
    if request.method == 'POST':
        form = ComplaintForm(request.POST, instance=complaint, user=request.user)
        if form.is_valid():
            # Capturar TODOS os valores antigos antes de salvar
            old_data = {
                'id_ra': complaint.id_ra,
                'cpf_cliente': complaint.cpf_cliente,
                'nome_cliente': complaint.nome_cliente,
                'sobrenome': complaint.sobrenome,
                'email_cliente': complaint.email_cliente,
                'telefone': complaint.telefone,
                'loja_cod': complaint.loja_cod,
                'origem_contato': complaint.origem_contato,
                'descricao': complaint.descricao,
                'status': complaint.status,
                'analista': complaint.analista,
                'data_reclamacao': complaint.data_reclamacao,
                'data_resposta': complaint.data_resposta,
                'nota_satisfacao': complaint.nota_satisfacao,
                'feedback_text': complaint.feedback_text,
            }
            
            # Salvar o formulário
            form.save()
            
            # Atualizar a instância do banco para comparar
            complaint.refresh_from_db()
            
            # Mapear nomes de campos para labels amigáveis
            field_labels = {
                'id_ra': 'ID RA',
                'cpf_cliente': 'CPF do Cliente',
                'nome_cliente': 'Nome do Cliente',
                'sobrenome': 'Sobrenome',
                'email_cliente': 'E-mail do Cliente',
                'telefone': 'Telefone',
                'loja_cod': 'Código da Loja',
                'origem_contato': 'Origem do Contato',
                'descricao': 'Descrição',
                'status': 'Status',
                'analista': 'Responsável',
                'data_reclamacao': 'Data da Reclamação',
                'data_resposta': 'Data de Resposta',
                'nota_satisfacao': 'Nota de Satisfação',
                'feedback_text': 'Feedback do Cliente',
            }
            
            # Criar atividades para cada campo alterado
            changes = []
            
            for field_name, old_value in old_data.items():
                new_value = getattr(complaint, field_name)
                
                # Comparar valores, tratando None, strings vazias e ForeignKey
                # Para ForeignKey, comparar os IDs
                if field_name == 'analista':
                    old_id = old_value.id if old_value else None
                    new_id = new_value.id if new_value else None
                    if old_id != new_id:
                        field_label = field_labels.get(field_name, field_name)
                        old_name = old_value.username if old_value else "Não atribuído"
                        new_name = new_value.username if new_value else "Não atribuído"
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_name}" para "{new_name}"',
                            tipo_interacao='atualizacao'
                        )
                        changes.append(field_name)
                    continue
                
                # Para outros campos, normalizar None e strings vazias
                old_val = old_value if old_value not in (None, "") else ""
                new_val = new_value if new_value not in (None, "") else ""
                
                if str(old_val).strip() != str(new_val).strip():
                    field_label = field_labels.get(field_name, field_name)
                    
                    # Formatação especial para alguns campos
                    if field_name == 'status':
                        old_display = dict(Complaint.STATUS_CHOICES).get(old_value, str(old_value))
                        new_display = complaint.get_status_display()
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_display}" para "{new_display}"',
                            tipo_interacao='mudanca_status'
                        )
                    elif field_name == 'origem_contato':
                        old_display = dict(Complaint.ORIGEM_CHOICES).get(old_value, str(old_value))
                        new_display = complaint.get_origem_contato_display()
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_display}" para "{new_display}"',
                            tipo_interacao='atualizacao'
                        )
                    elif field_name in ['data_reclamacao', 'data_resposta']:
                        old_str = old_value.strftime('%d/%m/%Y') if old_value else "Não informada"
                        new_str = new_value.strftime('%d/%m/%Y') if new_value else "Não informada"
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_str}" para "{new_str}"',
                            tipo_interacao='atualizacao'
                        )
                    else:
                        # Para outros campos, mostrar valores antigo e novo
                        old_str = str(old_value) if old_value else "(vazio)"
                        new_str = str(new_value) if new_value else "(vazio)"
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_str}" para "{new_str}"',
                            tipo_interacao='atualizacao'
                        )
                    changes.append(field_name)
            
            # Se não houver mudanças, criar atividade genérica
            if not changes:
                Activity.objects.create(
                    complaint=complaint,
                    usuario=request.user,
                    comentario='Informações da reclamação foram atualizadas',
                    tipo_interacao='atualizacao'
                )
            
            messages.success(request, 'Reclamação atualizada com sucesso!')
            return redirect('complaint_detail', pk=complaint.pk)
    else:
        form = ComplaintForm(instance=complaint, user=request.user)
    return render(request, 'core/complaint_form.html', {'form': form, 'complaint': complaint})

@login_required
def complaint_delete(request, pk):
    from django.contrib.auth import authenticate
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para excluir reclamações.')
        return redirect('complaint_list')
    
    complaint = get_object_or_404(Complaint, pk=pk)
    
    # Trava de segurança por departamento
    if not request.user.is_administrador():
        if complaint.department != request.user.department:
            messages.error(request, 'Você não tem permissão para excluir esta reclamação.')
            return redirect('dashboard')
    
    if request.method == 'POST':
        password = request.POST.get('password')
        
        if not password:
            messages.error(request, 'Por favor, informe sua senha para confirmar a exclusão.')
            return render(request, 'core/complaint_confirm_delete.html', {'complaint': complaint})
        
        # Verificar senha
        user = authenticate(request, username=request.user.username, password=password)
        if not user:
            messages.error(request, 'Senha incorreta. Tente novamente.')
            return render(request, 'core/complaint_confirm_delete.html', {'complaint': complaint})
        
        # Criar atividade antes de excluir
        Activity.objects.create(
            complaint=complaint,
            usuario=request.user,
            comentario=f'Reclamação {complaint.id_ra} foi excluída pelo gestor {request.user.username}',
            tipo_interacao='atualizacao'
        )
        
        AuditLog.objects.create(
            usuario=request.user,
            action='delete',
            target_type='Complaint',
            target_id=complaint.id,
            detalhes_json={'id_ra': complaint.id_ra}
        )
        complaint.delete()
        messages.success(request, 'Reclamação excluída com sucesso!')
        return redirect('complaint_list')
    
    return render(request, 'core/complaint_confirm_delete.html', {'complaint': complaint})

@login_required
def complaint_bulk_delete(request):
    """Exclusão em massa de reclamações - apenas para gestores"""
    from django.contrib.auth import authenticate
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para excluir reclamações.')
        return redirect('complaint_list')
    
    if request.method == 'POST':
        password = request.POST.get('password')
        delete_all = request.POST.get('delete_all') == 'on'
        selected_ids = request.POST.getlist('selected_complaints')
        
        if not password:
            messages.error(request, 'Por favor, informe sua senha para confirmar a exclusão.')
            return redirect('complaint_list')
        
        # Verificar senha
        user = authenticate(request, username=request.user.username, password=password)
        if not user:
            messages.error(request, 'Senha incorreta. Tente novamente.')
            return redirect('complaint_list')
        
        if delete_all:
            # Excluir todas as reclamações
            total = Complaint.objects.count()
            # Criar logs antes de excluir
            for complaint in Complaint.objects.all():
                AuditLog.objects.create(
                    usuario=request.user,
                    action='delete',
                    target_type='Complaint',
                    target_id=complaint.id,
                    detalhes_json={'id_ra': complaint.id_ra}
                )
            Complaint.objects.all().delete()
            messages.success(request, f'Todas as {total} reclamações foram excluídas com sucesso!')
        elif selected_ids:
            # Excluir selecionadas - pode vir como string separada por vírgula
            if isinstance(selected_ids, list) and len(selected_ids) > 0 and ',' in selected_ids[0]:
                selected_ids = selected_ids[0].split(',')
            # Converter para inteiros
            try:
                selected_ids = [int(id) for id in selected_ids if id]
            except (ValueError, TypeError):
                messages.error(request, 'IDs inválidos selecionados.')
                return redirect('complaint_list')
            
            complaints = Complaint.objects.filter(pk__in=selected_ids)
            count = complaints.count()
            for complaint in complaints:
                AuditLog.objects.create(
                    usuario=request.user,
                    action='delete',
                    target_type='Complaint',
                    target_id=complaint.id,
                    detalhes_json={'id_ra': complaint.id_ra}
                )
            complaints.delete()
            messages.success(request, f'{count} reclamação(ões) excluída(s) com sucesso!')
        else:
            messages.error(request, 'Nenhuma reclamação selecionada.')
        
        return redirect('complaint_list')
    
    return redirect('complaint_list')

@login_required
def import_complaints_xlsx(request):
    """Importar reclamações de arquivo XLSX"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para importar dados.')
        return redirect('dashboard')
    
    from .models import Department
    # Se for gestor, usa o depto dele. Se for admin, tenta pegar do POST ou padrão CS Clientes
    target_dept = request.user.department
    if request.user.is_administrador():
        dept_id = request.POST.get('department')
        if dept_id:
            target_dept = Department.objects.filter(id=dept_id).first()
        if not target_dept:
            target_dept = Department.objects.filter(slug='cs-clientes').first()

    if request.method == 'POST':
        if 'xlsx_file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo XLSX.')
            return render(request, 'core/import_complaints.html')
        
        file = request.FILES['xlsx_file']
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Por favor, selecione um arquivo XLSX válido.')
            return render(request, 'core/import_complaints.html')
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            imported = 0
            updated = 0
            skipped = 0
            errors = []
            total_rows = 0
            
            # Contar total de linhas (exceto cabeçalho)
            for row in ws.iter_rows(min_row=2, values_only=True):
                total_rows += 1
            
            # Começar da linha 2 (linha 1 é cabeçalho)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                status, msg = _process_complaint_row(row, row_num, target_dept)
                
                if status == 'created':
                    imported += 1
                elif status == 'updated':
                    updated += 1
                elif status == 'skipped':
                    skipped += 1
                    errors.append(msg)
                elif status == 'error':
                    errors.append(msg)
            
            # Mensagens
            total_processed = imported + updated
            if total_processed > 0:
                msg = f"Importação concluída! Processadas {total_rows} linha(s) da planilha. "
                msg += f"{imported} reclamação(ões) criada(s), {updated} atualizada(s)."
                if skipped > 0:
                    msg += f" {skipped} linha(s) ignorada(s) (ID RA vazio)."
                if len(errors) > skipped:
                    msg += f" {len(errors) - skipped} aviso(s)."
                messages.success(request, msg)
            else:
                messages.error(request, f"Nenhuma reclamação foi importada. Verifique os erros abaixo.")
            
            if errors:
                for error in errors[:15]:
                    messages.warning(request, error)
                if len(errors) > 15:
                    messages.warning(request, f"... e mais {len(errors) - 15} aviso(s)/erro(s).")
            
            return redirect('complaint_list')
            
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
            return render(request, 'core/import_complaints.html')
    
    return render(request, 'core/import_complaints.html')

def _process_complaint_row(row, row_num, target_dept):
    """
    Processa uma linha da planilha de importação.
    Retorna (status, mensagem)
    status: 'created', 'updated', 'skipped', 'error'
    """
    try:
        # Mapeamento de status e tipos
        status_map = {
            'pendente': 'pendente', 'em andamento': 'em_andamento', 'em réplica': 'em_replica',
            'aguardando avaliação': 'aguardando_avaliacao', 'resolvido': 'resolvido', 'resolvida': 'resolvido',
        }
        tipo_map = {
            'nota fiscal': 'nota_fiscal', 'pagamento não processado - cartão': 'pagamento_cartao',
            'pagamento não processado - pix': 'pagamento_pix', 'pagamento não processado - checkout web': 'pagamento_checkout',
            'assinatura mensal': 'assinatura_mensal', 'lavagem': 'lavagem', 'secagem': 'secagem',
            'atendimento': 'atendimento', 'sistema/totem': 'sistema_totem', 'totem': 'sistema_totem',
            'cupons': 'cupons', 'outros': 'outros',
        }
        volta_negocio_map = {'sim': 'sim', 's': 'sim', 'não': 'nao', 'nao': 'nao', 'n': 'nao'}

        # Mapear colunas
        loja_cod = str(row[0]).strip() if len(row) > 0 and row[0] else 'Não informado'
        nome_completo = str(row[1]).strip() if len(row) > 1 and row[1] else 'Nome não informado'
        id_ra = str(row[2]).strip() if len(row) > 2 and row[2] else None
        
        # Validação ID RA
        if not id_ra or id_ra == '':
            return 'skipped', f"Linha {row_num}: ID RA está vazio - linha ignorada"

        cpf = str(row[3]).strip() if len(row) > 3 and row[3] else None
        email_cliente = str(row[4]).strip() if len(row) > 4 and row[4] else None
        telefone = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        data_reclamacao = row[6] if len(row) > 6 and row[6] else None
        problema = str(row[7]).strip().lower() if len(row) > 7 and row[7] else None
        status = str(row[8]).strip().lower() if len(row) > 8 and row[8] else 'pendente'
        analista_nome = str(row[9]).strip() if len(row) > 9 and row[9] else None
        nota = row[10] if len(row) > 10 and row[10] else None
        volta_negocio = str(row[11]).strip().lower() if len(row) > 11 and row[11] else None

        # Processar CPF
        cpf_clean = re.sub(r'\D', '', str(cpf)) if cpf else '00000000000'
        if len(cpf_clean) != 11: cpf_clean = '00000000000'

        # Dividir nome
        nome_parts = str(nome_completo).split(maxsplit=1)
        nome_cliente = nome_parts[0] if nome_parts else 'Nome não informado'
        sobrenome = nome_parts[1] if len(nome_parts) > 1 else ''

        # Processar Data
        data_reclamacao_value = timezone.now().date()
        if data_reclamacao:
            if isinstance(data_reclamacao, datetime):
                data_reclamacao_value = data_reclamacao.date()
            elif isinstance(data_reclamacao, str) and data_reclamacao.strip():
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                    try:
                        data_reclamacao_value = datetime.strptime(data_reclamacao.strip(), fmt).date()
                        break
                    except ValueError: pass

        # Mapear Status e Tipo
        status_value = status_map.get(status, 'pendente')
        tipo_reclamacao_value = tipo_map.get(problema, 'outros') if problema else None
        volta_negocio_value = volta_negocio_map.get(volta_negocio, 'nao_informado') if volta_negocio else None

        # Buscar Analista
        analista_obj = None
        if analista_nome and analista_nome.lower().strip() not in ['selecione um analista (opcional)', 'não atribuido', '', 'nao atribuido']:
            try:
                parts = str(analista_nome).strip().split()
                q = Q()
                for p in parts:
                    q |= Q(first_name__icontains=p) | Q(last_name__icontains=p)
                analista_obj = User.objects.filter(role='analista', ativo=True).filter(q).first()
            except: pass

        # Nota
        nota_value = None
        if nota is not None:
            try:
                nota_value = max(0, min(10, int(float(nota))))
            except: pass

        # Email fallback
        if not email_cliente or '@' not in email_cliente:
            email_cliente = f'{cpf_clean}@importado.com'

        descricao = f'Importado da planilha' + (f' - Tipo: {problema}' if problema else ' - Tipo: Não informado')

        complaint, created = Complaint.objects.update_or_create(
            id_ra=str(id_ra).strip(),
            defaults={
                'cpf_cliente': cpf_clean,
                'nome_cliente': nome_cliente,
                'sobrenome': sobrenome,
                'email_cliente': email_cliente,
                'telefone': telefone,
                'loja_cod': loja_cod,
                'origem_contato': 'RA',
                'descricao': descricao,
                'status': status_value,
                'analista': analista_obj,
                'data_reclamacao': data_reclamacao_value,
                'tipo_reclamacao': tipo_reclamacao_value,
                'nota_satisfacao': nota_value,
                'volta_fazer_negocio': volta_negocio_value,
                'department': target_dept,
            }
        )
        return ('created' if created else 'updated'), None

    except Exception as e:
        return 'error', f"Linha {row_num}: {str(e)}"

@login_required
def import_complaints_batch(request):
    """API para importação em lote (Batch)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
        
    if not (request.user.is_gestor() or request.user.is_administrador()):
        return JsonResponse({'error': 'Permissão negada'}, status=403)

    try:
        import json
        data = json.loads(request.body)
        rows = data.get('rows', [])
        
        # Determinar departamento (mesma lógica do import normal)
        from .models import Department
        target_dept = request.user.department
        if request.user.is_administrador():
            # Tentar pegar dept do body se enviado, ou usar padrão
             target_dept = Department.objects.filter(slug='cs-clientes').first()

        results = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }

        for i, row in enumerate(rows):
            db_status, msg = _process_complaint_row(row, i, target_dept)
            
            if db_status == 'created': results['created'] += 1
            elif db_status == 'updated': results['updated'] += 1
            elif db_status == 'skipped': results['skipped'] += 1
            elif db_status == 'error': results['errors'].append(msg)

        return JsonResponse(results)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

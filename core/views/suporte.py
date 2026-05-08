from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg, Exists, OuterRef
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from datetime import datetime, timedelta
import json
from ..models import (
    Turno, AnalistaEscala, FolgaManual, EscalaRascunho, Department, 
    IndicadorDesempenho, MetaMensalGlobal, KanbanBoard, 
    KanbanList, CardLabel, Store, StoreAudit, 
    StoreAuditIssue, StoreAuditItem, SystemNotification,
    User, DailyAuditQuota, ModeloEscala, ConfiguracaoEscala
)
from ..forms import StoreForm

@login_required
def sites_view(request):
    """Página de Sites e Sistemas - NRS Suporte"""
    if not request.user.is_administrador():
        if not request.user.department or request.user.department.name != 'NRS Suporte':
            messages.error(request, 'Você não tem permissão para acessar as ferramentas de NRS Suporte.')
            return redirect('dashboard')
            
    return render(request, 'core/sites.html')

@login_required
def localizacao_view(request):
    """Página de Localização das Lojas - NRS Suporte"""
    if not request.user.is_administrador():
        if not request.user.department or request.user.department.name != 'NRS Suporte':
            messages.error(request, 'Você não tem permissão para acessar as ferramentas de NRS Suporte.')
            return redirect('dashboard')
            
    return render(request, 'core/localizacao.html')

@login_required
def escala_view(request):
    """Página de Escala - NRS Suporte"""
    user = request.user
    if not user.is_administrador():
        if not user.department or user.department.name not in ['NRS Suporte', 'RH']:
            messages.error(request, 'Você não tem permissão para acessar as ferramentas de NRS Suporte.')
            return redirect('dashboard')
    
    rascunho_id = request.GET.get('rascunho_id')
    rascunho_obj = None
    if rascunho_id:
        rascunho_obj = get_object_or_404(EscalaRascunho, id=rascunho_id)
        turnos = Turno.objects.filter(ativo=True, rascunho=rascunho_obj).order_by('ordem', 'nome')
        analistas = AnalistaEscala.objects.filter(ativo=True, rascunho=rascunho_obj).select_related('turno', 'modelo_escala').order_by('turno__ordem', 'ordem', 'nome')
        folgas = FolgaManual.objects.filter(rascunho=rascunho_obj).select_related('analista')
    else:
        turnos = Turno.objects.filter(ativo=True, rascunho__isnull=True).order_by('ordem', 'nome')
        analistas = AnalistaEscala.objects.filter(ativo=True, rascunho__isnull=True).select_related('turno', 'modelo_escala').order_by('turno__ordem', 'ordem', 'nome')
        folgas = FolgaManual.objects.filter(rascunho__isnull=True).select_related('analista')
    
    turnos_data = [{
        'id': str(t.id),
        'nome': t.nome,
        'horario': t.horario,
        'cor': t.cor,
        'ordem': t.ordem
    } for t in turnos]
    
    analistas_data = [{
        'id': str(a.id),
        'nome': a.nome,
        'turno': a.turno.nome if a.turno else None,
        'turno_id': a.turno.id if a.turno else None,
        'modelo_escala_id': a.modelo_escala.id if a.modelo_escala else None,
        'pausa': a.pausa,
        'data_primeira_folga': a.data_primeira_folga.isoformat() if a.data_primeira_folga else None,
        'ordem': a.ordem
    } for a in analistas]
    
    folgas_data = {}
    for f in folgas:
        key = f"{f.analista.id}-{f.data.year}-{f.data.month}-{f.data.day}"
        folgas_data[key] = {
            'id': f.id,
            'tipo': f.tipo,
            'motivo': f.motivo
        }
    
    # RH vê a escala em modo somente-leitura (sem botões de editar/adicionar analista/turno)
    is_rh = user.department and user.department.name == 'RH'
    is_admin = (user.is_gestor() or user.is_administrador()) and not is_rh
    can_export = True  # Todos os usuários logados com acesso à escala podem ver Resumo e Exportar
    
    is_planejamento_mode = request.resolver_match.url_name == 'planejamento_escala'
    is_config_mode = request.resolver_match.url_name == 'configuracao_escalas'

    # Segurança: Apenas gestores e administradores acessam planejamento e configuração
    if (is_planejamento_mode or is_config_mode) and not (user.is_gestor() or user.is_administrador()):
        messages.error(request, 'Você não tem permissão para acessar esta área da escala.')
        return redirect('escala')

    modelos = ModeloEscala.objects.all().order_by('nome')
    modelos_data = [{
        'id': m.id,
        'nome': m.nome,
        'dias_trabalhados': m.dias_trabalhados,
        'dias_folga': m.dias_folga,
        'tipo': m.tipo,
        'permite_fim_de_semana': m.permite_fim_de_semana,
        'observacao': m.observacao
    } for m in modelos]
    
    config = ConfiguracaoEscala.objects.first()
    modelo_principal_id = config.modelo_escala_principal.id if config and config.modelo_escala_principal else None

    user_analista = None
    if user.is_authenticated:
        user_analista = AnalistaEscala.objects.filter(user=user, rascunho__isnull=True, ativo=True).first()

    context = {
        'turnos_json': json.dumps(turnos_data),
        'analistas_json': json.dumps(analistas_data),
        'folgas_json': json.dumps(folgas_data),
        'modelos_json': json.dumps(modelos_data),
        'modelo_principal_id': modelo_principal_id,
        'is_admin': is_admin,
        'is_admin_json': 'true' if is_admin else 'false',
        'can_export': can_export,
        'is_rascunho': rascunho_obj is not None,
        'rascunho_id': str(rascunho_obj.id) if rascunho_obj else None,
        'rascunho_modelo_id': str(rascunho_obj.modelo_escala.id) if rascunho_obj and rascunho_obj.modelo_escala else None,
        'rascunho_nome': rascunho_obj.nome if rascunho_obj else None,
        'is_planejamento_mode': is_planejamento_mode,
        'is_config_mode': is_config_mode,
        'user_analista_id': str(user_analista.id) if user_analista else None,
    }
    
    return render(request, 'core/escala.html', context)

@login_required
def calendar_view(request):
    """Visualização do calendário"""
    user = request.user
    is_manager = user.role in ['gestor', 'administrador']
    is_nrs_analyst = user.role == 'analista' and user.department and user.department.name == 'NRS Suporte'
    
    can_create = is_manager or is_nrs_analyst
    
    context = {
        'can_edit': can_create,
        'user_id': user.id,
        'user_full_name': f"{user.first_name} {user.last_name}".strip() or user.username,
        'is_manager': is_manager,
    }
    return render(request, 'core/calendar.html', context)

@login_required
def knowledge_base_view(request):
    """Visualização da Base de Conhecimento"""
    can_edit = request.user.is_administrador() or request.user.is_gestor()
    return render(request, 'core/knowledge_base.html', {'can_edit': can_edit})

@login_required
def performance_view(request):
    """Página de Desempenho do Time"""
    user = request.user
    nrs_dept = Department.objects.filter(name='NRS Suporte').first()
    if not nrs_dept:
        nrs_dept = Department.objects.filter(slug='nrs-suporte').first()
    if not nrs_dept and user.department:
        nrs_dept = user.department
    if not nrs_dept:
        nrs_dept = Department.objects.first()
    
    is_analista = user.role == 'analista'
    can_edit = user.role in ['gestor', 'administrador']
    page_title = "Meu Desempenho" if is_analista else "Desempenho do Time"

    if can_edit and nrs_dept:
        analistas = User.objects.filter(
            Q(department=nrs_dept) | 
            Q(indicadores_desempenho__department=nrs_dept)
        ).distinct().order_by('first_name', 'username')
    else:
        analistas = []
    
    selected_analista_id = request.GET.get('analista_id')
    if is_analista:
        selected_analista_id = user.id
    elif not selected_analista_id and analistas:
        selected_analista_id = analistas.first().id if analistas.exists() else None
    
    if selected_analista_id and str(selected_analista_id).isdigit():
        selected_analista_id = int(selected_analista_id)

    period = request.GET.get('period', '6')
    try:
        period_int = int(period) if period != 'all' else None
    except ValueError:
        period_int = 6

    kpis_data = []
    if selected_analista_id and nrs_dept:
        kpis_query = IndicadorDesempenho.objects.filter(
            analista_id=selected_analista_id,
            department=nrs_dept
        ).order_by('ano', 'mes')
        
        kpis = list(kpis_query)
        if period_int and len(kpis) > period_int:
            kpis = kpis[-period_int:]
        
        metas_globais = {
            f"{m.mes:02d}/{m.ano}": m 
            for m in MetaMensalGlobal.objects.filter(department=nrs_dept)
        }
        
        for kpi in kpis:
            label = f"{kpi.mes:02d}/{kpi.ano}"
            meta = metas_globais.get(label)
            kpis_data.append({
                'id': kpi.id,
                'mes': kpi.mes,
                'ano': kpi.ano,
                'label': label,
                'nps': float(kpi.nps) if kpi.nps else None,
                'tme': kpi.tme,
                'chats': kpi.chats,
                'meta_tme': meta.meta_tme if meta else None,
                'meta_nps': float(meta.meta_nps) if meta and meta.meta_nps else None,
                'meta_chats': meta.meta_chats if meta else None,
            })
    
    selected_analista = None
    if selected_analista_id:
        try:
            selected_analista = User.objects.get(id=selected_analista_id)
        except User.DoesNotExist:
            pass

    analistas_list = [{
        'id': a.id,
        'name': a.get_full_name() or a.username,
        'is_selected': a.id == selected_analista_id
    } for a in analistas]

    context = {
        'page_title': page_title,
        'is_analista': is_analista,
        'can_edit': can_edit,
        'can_edit_str': "true" if can_edit else "false",
        'analistas_list': analistas_list,
        'selected_analista_id': selected_analista_id,
        'selected_analista': selected_analista,
        'kpis_json': json.dumps(kpis_data),
        'current_period': period,
        'analistas_json': json.dumps([{
            'id': a.id,
            'nome': a.get_full_name() or a.username
        } for a in analistas]),
    }
    return render(request, 'core/desempenho.html', context)

@login_required
def quadro_view(request):
    """Visualização do Quadro Kanban"""
    board = KanbanBoard.objects.filter(owner=request.user).first()
    if not board:
        board = KanbanBoard.objects.first() or KanbanBoard.objects.create(
            name='Quadro Principal', owner=request.user, background_color='#2563eb'
        )
    
    if not board.lists.exists():
        KanbanList.objects.create(board=board, name='A Fazer', position=0)
        KanbanList.objects.create(board=board, name='Em Andamento', position=1)
        KanbanList.objects.create(board=board, name='Concluído', position=2)
    
    if not board.labels.exists():
        default_labels = [
            ('Urgente', '#ef4444'), ('Importante', '#f97316'), ('Normal', '#6b7280'),
            ('Baixa', '#3b82f6'), ('Reunião', '#6366f1'), ('Documentação', '#0891b2'),
            ('Bug', '#dc2626'), ('Feature', '#059669'), ('Concluído', '#22c55e'),
        ]
        for name, color in default_labels:
            CardLabel.objects.create(board=board, name=name, color=color)
    
    listas = board.lists.filter(is_archived=False).prefetch_related('cards__labels').order_by('position')
    labels_data = [{'id': l.id, 'name': l.name, 'color': l.color} for l in board.labels.all()]
    
    nrs_dept = Department.objects.filter(name='NRS Suporte').first()
    members = []
    if nrs_dept:
        nrs_users = User.objects.filter(department=nrs_dept, ativo=True).order_by('first_name', 'username')
        members = [{
            'id': u.id, 'name': u.get_full_name() or u.username,
            'initials': (u.first_name[:1] + u.last_name[:1]).upper() if u.first_name and u.last_name else u.username[:2].upper(),
            'role': u.get_role_display()
        } for u in nrs_users]
    
    return render(request, 'core/quadro.html', {
        'board': board, 'listas': listas, 'labels_json': json.dumps(labels_data), 'members_json': json.dumps(members),
    })

@login_required
def tasks_view(request):
    """View para a aba de tarefas e rotina"""
    user = request.user
    is_manager = user.role in ['gestor', 'administrador']
    is_nrs_analyst = (user.role == 'analista' and user.department and user.department.name == 'NRS Suporte')
    show_create_button = is_manager or is_nrs_analyst or user.is_administrador()
    return render(request, 'core/tarefas.html', {
        'title': 'Tarefas e Solicitações', 'is_manager': is_manager, 'show_create_button': show_create_button,
    })

@login_required
def solicitacoes_view(request):
    """View para a aba de solicitações de estorno (CS Clientes)"""
    user = request.user
    is_manager = user.role in ['gestor', 'administrador']
    is_cs_clientes = user.department and user.department.name == 'CS Clientes'
    return render(request, 'core/solicitacoes.html', {
        'title': 'Solicitações de Estorno', 'is_manager': is_manager, 'is_cs_clientes': is_cs_clientes,
    })

@login_required
@ensure_csrf_cookie
def verificacao_lojas(request):
    """Página principal de verificação de lojas (NRS Suporte)"""
    tab = request.GET.get('tab', 'lojas')
    if tab == 'all': tab = 'lojas'
    scope = request.GET.get('scope', 'all')
    search_query = request.GET.get('q', '')
    
    from ..models import AnalystAssignment
    my_ids = []
    if scope == 'my_stores' and request.user.is_authenticated:
        my_ids = list(AnalystAssignment.objects.filter(analyst=request.user, active=True).values_list('store_id', flat=True))

    has_open_issue = StoreAuditIssue.objects.filter(store=OuterRef('pk'), status='aberta')
    has_audit = StoreAudit.objects.filter(store=OuterRef('pk'))

    stores_queryset = Store.objects.annotate(
        has_open_issue=Exists(has_open_issue),
        has_audit=Exists(has_audit)
    ).order_by('code')

    from django.core.cache import cache
    cache_key = f'auditoria_stats_{scope}_{request.user.id}'
    cached_stats = cache.get(cache_key)
    
    if cached_stats:
        total_active = cached_stats['total_active']
        irregular_count = cached_stats['irregular_count']
        ok_count = cached_stats['ok_count']
        suspended_count = cached_stats['suspended_count']
        irregular_store_ids = cached_stats['irregular_store_ids']
        weekly_audits_count = cached_stats['weekly_audits_count']
        weekly_irregularities_count = cached_stats['weekly_irregularities_count']
        compliance_rate = cached_stats['compliance_rate']
    else:
        # Preamble for stats
        start_week = timezone.now() - timedelta(days=timezone.now().weekday())
        start_week = start_week.replace(hour=0, minute=0, second=0, microsecond=0)

        base_stats_query = Store.objects.filter(active=True)
        if scope == 'my_stores' and request.user.is_authenticated:
            base_stats_query = base_stats_query.filter(id__in=my_ids)

        total_active = base_stats_query.count()
        
        irregular_qs = StoreAuditIssue.objects.filter(status='aberta')
        if scope == 'my_stores' and request.user.is_authenticated:
            irregular_qs = irregular_qs.filter(store_id__in=my_ids)
        
        irregular_store_ids = set(irregular_qs.values_list('store_id', flat=True).distinct())
        irregular_count = len(irregular_store_ids)
        
        # Optimized ok_count using Exists
        has_audit_any = StoreAudit.objects.filter(store=OuterRef('pk'))
        has_issue_any = StoreAuditIssue.objects.filter(store=OuterRef('pk'), status='aberta')
        ok_count = base_stats_query.filter(Exists(has_audit_any)).exclude(Exists(has_issue_any)).count()
        
        suspended_query = Store.objects.filter(active=False)
        if scope == 'my_stores' and request.user.is_authenticated:
             suspended_query = suspended_query.filter(id__in=my_ids)
        suspended_count = suspended_query.count()

        # Weekly Stats Aggregation
        weekly_audits_qs = StoreAudit.objects.filter(created_at__gte=start_week)
        weekly_total_qs = StoreAuditItem.objects.filter(audit__created_at__gte=start_week)
        
        if scope == 'my_stores' and request.user.is_authenticated:
            weekly_audits_qs = weekly_audits_qs.filter(analyst=request.user)
            weekly_total_qs = weekly_total_qs.filter(audit__analyst=request.user)
            
        weekly_audits_count = weekly_audits_qs.count()
        
        # Single aggregate query for items
        weekly_stats = weekly_total_qs.aggregate(
            total=Count('id'),
            compliant=Count('id', filter=Q(is_compliant=True)),
            irregular=Count('id', filter=Q(is_compliant=False))
        )
        
        total_items = weekly_stats['total']
        weekly_irregularities_count = weekly_stats['irregular']
        compliance_rate = (weekly_stats['compliant'] / total_items * 100) if total_items > 0 else 100
        
        cache.set(cache_key, {
            'total_active': total_active, 
            'irregular_count': irregular_count, 
            'ok_count': ok_count,
            'suspended_count': suspended_count, 
            'irregular_store_ids': irregular_store_ids,
            'weekly_audits_count': weekly_audits_count,
            'weekly_irregularities_count': weekly_irregularities_count,
            'compliance_rate': compliance_rate
        }, 300)

    if search_query:
        stores_queryset = stores_queryset.filter(Q(code__icontains=search_query) | Q(city__icontains=search_query))
        tab = 'all'
    elif tab == 'suspended': stores_queryset = stores_queryset.filter(active=False)
    elif tab == 'irregular': stores_queryset = stores_queryset.filter(active=True, id__in=irregular_store_ids)
    elif tab == 'verified': stores_queryset = stores_queryset.filter(active=True, audits__isnull=False).exclude(id__in=irregular_store_ids).distinct()
    else: stores_queryset = stores_queryset.filter(active=True)

    if scope == 'my_stores' and request.user.is_authenticated:
        stores_queryset = stores_queryset.filter(id__in=my_ids)

    history, stores, pending_issues, pending_issues_page, pending_issues_count = [], [], [], None, 0

    if tab == 'history':
        history_qs = StoreAudit.objects.select_related('store', 'analyst').order_by('-created_at')
        if scope == 'my_stores' and request.user.is_authenticated:
            history_qs = history_qs.filter(analyst=request.user)
        paginator = Paginator(history_qs, 50)
        history = paginator.get_page(request.GET.get('page'))
    elif tab == 'irregular' and request.user.role in ['gestor', 'administrador', 'analista']:
        # Note: Added 'analista' to irregular if needed, but per logic manager sees pending issues
        pending_qs = StoreAuditIssue.objects.filter(status='aberta').select_related('store').prefetch_related('items', 'items__audit__analyst').order_by('-created_at')
        if scope == 'my_stores' and request.user.role == 'analista':
            pending_qs = pending_qs.filter(store_id__in=my_ids)
            
        pending_paginator = Paginator(pending_qs, 25)
        pending_issues_page = pending_paginator.get_page(request.GET.get('pending_page', 1))
        pending_issues, pending_issues_count = pending_issues_page.object_list, pending_paginator.count
    else:
        paginator = Paginator(stores_queryset, 25)
        stores = paginator.get_page(request.GET.get('page'))
    
    if stores:
        page_store_ids = [s.id for s in stores]
        start_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        monthly_counts = {item['store_id']: item['count'] for item in StoreAudit.objects.filter(store_id__in=page_store_ids, created_at__gte=start_month).values('store_id').annotate(count=Count('id'))}
        
        latest_audits = {}
        for audit in StoreAudit.objects.filter(store_id__in=page_store_ids).select_related('analyst').order_by('store_id', '-created_at'):
            if audit.store_id not in latest_audits: latest_audits[audit.store_id] = audit

        stores_with_audits = set(StoreAudit.objects.filter(store_id__in=page_store_ids).values_list('store_id', flat=True).distinct())

        for s in stores:
            s.latest_audit = latest_audits.get(s.id)
            s.audits_this_month_count = monthly_counts.get(s.id, 0)
            _has_open_issue, _has_audit = s.id in irregular_store_ids, s.id in stores_with_audits
            if not s.active: s.ui_status = 'suspended'
            elif _has_open_issue: s.ui_status = 'irregular'
            elif _has_audit: s.ui_status = 'compliant'
            else: s.ui_status = 'pending'

    return render(request, 'core/verificacao_lojas.html', {
        'title': 'Auditoria de Lojas', 'stores': stores, 'history': history, 'total_stores': total_active,
        'scope': scope, 'tab': tab, 'verified_count': ok_count, 'ok_count': ok_count, 'irregular_count': irregular_count,
        'suspended_count': suspended_count, 'pending_issues': pending_issues, 'pending_issues_page': pending_issues_page,
        'irregular_store_ids': irregular_store_ids, 'play_sound': request.session.pop('play_irregularity_sound', False),
        'search_query': search_query, 'pending_issues_count': pending_issues_count,
        'weekly_audits_count': weekly_audits_count, 'weekly_irregularities_count': weekly_irregularities_count,
        'compliance_rate': round(compliance_rate, 1),
    })

@login_required
def api_store_detail(request, store_id):
    try:
        store = Store.objects.get(id=store_id)
        last_audit = StoreAudit.objects.filter(store=store).select_related('analyst').order_by('-created_at').first()
        return JsonResponse({
            'success': True,
            'store': {
                'id': store.id, 'code': store.code, 'city': store.city or 'Não informada',
                'active': store.active, 'needs_reverification': store.needs_reverification,
                'last_audit_date': timezone.localtime(last_audit.created_at).strftime('%d/%m/%Y %H:%M') if last_audit else None,
                'last_audit_result': store.last_audit_result,
                'analyst_name': last_audit.analyst.get_full_name() or last_audit.analyst.username if last_audit else None,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def store_audit_create(request, store_id):
    store = get_object_or_404(Store, id=store_id)
    if request.method == 'POST':
        if request.user.role == 'analista':
            quota = DailyAuditQuota.get_or_create_today(request.user)
            if quota.is_quota_reached:
                messages.error(request, f'⚠️ Limite diário atingido! ({quota.audits_completed}/{quota.daily_quota})')
                return redirect('verificacao_lojas')
        
        audit = StoreAudit.objects.create(analyst=request.user, store=store)
        items_slugs = ['cameras', 'estofados', 'cestos_medidas', 'layout', 'tv', 'totem', 'limpeza', 'marketing']
        has_irregularity = False
        
        for slug in items_slugs:
            status = request.POST.get(f'status_{slug}')
            is_compliant = (status == 'conformidade')
            rec_val = request.POST.get('cameras_recording') if slug == 'cameras' else None
            
            audit_item = StoreAuditItem.objects.create(
                audit=audit, item_name=slug, is_compliant=is_compliant, photo=request.FILES.get(f'photo_{slug}'),
                description=request.POST.get(f'desc_{slug}', ''),
                cameras_recording=(rec_val == 'yes') if rec_val else None,
                cameras_recording_mode=request.POST.get('cameras_mode') if rec_val == 'yes' else None
            )
            
            if not is_compliant:
                has_irregularity = True
                issue, _ = StoreAuditIssue.objects.get_or_create(store=store, status='aberta')
                audit_item.issue = issue
                audit_item.save()
        
        store.last_audit_date, store.last_audit_result, store.needs_reverification = timezone.now(), ('irregular' if has_irregularity else 'conforme'), False
        store.save()
        
        if request.user.role == 'analista':
            if not DailyAuditQuota.get_or_create_today(request.user).increment_audits():
                messages.warning(request, '✅ Meta diária concluída!')
        
        if has_irregularity:
            messages.warning(request, f"Auditoria da loja {store.code} finalizada com irregularidades.")
            request.session['play_irregularity_sound'] = True
        else:
            messages.success(request, f"Auditoria da loja {store.code} finalizada com sucesso.")
        return redirect('/verificacao-lojas/?tab=management')

    history = StoreAudit.objects.filter(store=store).order_by('-created_at')[:10]
    return render(request, 'core/store_audit_form.html', {
        'store': store, 'history': history, 'items_choices': [
            ('cameras', 'Câmeras'), ('estofados', 'Estofados'), ('cestos_medidas', 'Cestos de medidas'),
            ('layout', 'Layout'), ('tv', 'TV'), ('totem', 'Totem'), ('limpeza', 'Limpeza da loja'), ('marketing', 'Marketing'),
        ], 'title': f'Auditoria: Loja {store.code}', 'last_audit': history.first(),
        'last_audit_items': history.first().items.filter(is_compliant=False) if history.exists() else [],
        'needs_reverification': store.needs_reverification,
    })

@login_required
def store_issue_resolve(request, issue_id):
    if not (request.user.role in ['gestor', 'administrador']):
        messages.error(request, "Acesso negado.")
        return redirect('verificacao_lojas')
    issue = get_object_or_404(StoreAuditIssue, id=issue_id)
    if request.method == 'POST':
        notes, stage, channel = request.POST.get('gestor_notes', ''), request.POST.get('resolution_stage', ''), request.POST.get('notification_channel', '')
        if not channel:
            messages.error(request, "Informe o canal de notificação.")
            return redirect('verificacao_lojas')
        
        history = issue.resolution_history if isinstance(issue.resolution_history, list) else []
        history.append({'timestamp': timezone.now().isoformat(), 'user': request.user.username, 'stage': stage, 'channel': channel, 'notes': notes})
        issue.resolution_history, issue.gestor_notes, issue.resolution_stage, issue.notification_channel = history, notes, stage, channel
        issue.status, issue.resolved_at, issue.resolved_by = 'resolvida', timezone.now(), request.user
        issue.save()
        messages.success(request, f"Pendência em {issue.store.code} resolvida.")
    return redirect('verificacao_lojas')

@login_required
def store_create(request):
    if request.user.role not in ['gestor', 'administrador']: return redirect('verificacao_lojas')
    if request.method == 'POST':
        form = StoreForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Loja criada.")
            return redirect('verificacao_lojas')
    return render(request, 'core/store_form.html', {'form': StoreForm(), 'title': 'Cadastrar Nova Loja'})

@login_required
def store_edit(request, store_id):
    if request.user.role not in ['gestor', 'administrador']: return redirect('verificacao_lojas')
    store = get_object_or_404(Store, id=store_id)
    if request.method == 'POST':
        form = StoreForm(request.POST, instance=store)
        if form.is_valid():
            form.save()
            messages.success(request, f"Loja {store.code} atualizada.")
            return redirect('verificacao_lojas')
    return render(request, 'core/store_form.html', {'form': StoreForm(instance=store), 'title': f'Editar Loja: {store.code}'})

@login_required
def store_delete(request, store_id):
    if request.user.role not in ['gestor', 'administrador']: return redirect('verificacao_lojas')
    store = get_object_or_404(Store, id=store_id)
    if request.method == 'POST':
        code = store.code
        store.delete()
        messages.success(request, f"Loja {code} excluída.")
        return redirect('verificacao_lojas')
    return render(request, 'core/confirm_delete.html', {
        'title': f'Excluir Loja {store.code}', 'message': f'Confirmar exclusão da loja {store.code}?', 'back_url': 'verificacao_lojas'
    })

@login_required
def store_bulk_delete(request):
    if not request.user.is_administrador(): return redirect('verificacao_lojas')
    if request.method == 'POST':
        count = Store.objects.all().count()
        Store.objects.all().delete()
        messages.success(request, f"{count} lojas excluídas.")
        return redirect('verificacao_lojas')
    return render(request, 'core/confirm_delete.html', {
        'title': 'EXCLUIR TODAS AS LOJAS', 'message': 'EXCLUIR TODAS AS LOJAS E HISTÓRICOS?', 'back_url': 'verificacao_lojas'
    })

@login_required
def store_issue_edit(request, issue_id):
    if request.user.role not in ['gestor', 'administrador']: return redirect('verificacao_lojas')
    issue = get_object_or_404(StoreAuditIssue, id=issue_id)
    if request.method == 'POST':
        issue.status = request.POST.get('status', issue.status)
        issue.gestor_notes = request.POST.get('gestor_notes', issue.gestor_notes)
        issue.save()
        messages.success(request, "Pendência atualizada.")
        return redirect('verificacao_lojas')
    return render(request, 'core/issue_edit_form.html', {'issue': issue, 'title': f'Editar Pendência: {issue.store.code if issue.store else "???"}'})

@login_required
def store_issue_delete(request, issue_id):
    if request.user.role not in ['gestor', 'administrador']: return redirect('verificacao_lojas')
    issue = get_object_or_404(StoreAuditIssue, id=issue_id)
    if request.method == 'POST':
        issue.delete()
        messages.success(request, "Pendência excluída.")
        return redirect('verificacao_lojas')
    return render(request, 'core/confirm_delete.html', {
        'title': 'Excluir Pendência', 'message': f'Excluir pendência da loja {issue.store.code if issue.store else "???"}?', 'back_url': 'verificacao_lojas'
    })

@login_required
def import_stores_xlsx(request):
    if request.user.role not in ['gestor', 'administrador']: return redirect('verificacao_lojas')
    if request.method == 'POST':
        file = request.FILES.get('xlsx_file')
        if not file: return redirect('import_stores_xlsx')
        try:
            from openpyxl import load_workbook
            ws, created, updated = load_workbook(file, data_only=True).active, 0, 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip().upper() if row and row[0] else None
                if code:
                    _, is_created = Store.objects.get_or_create(code=code, defaults={'active': True})
                    if is_created: created += 1
                    else: updated += 1
            messages.success(request, f"Importação: {created} criadas, {updated} atualizadas.")
            return redirect('verificacao_lojas')
        except Exception as e:
            messages.error(request, f"Erro: {str(e)}")
    return render(request, 'core/import_stores.html', {'title': 'Importar Lojas (XLSX)'})

@login_required
def auditoria_atendimentos_view(request):
    if not (request.user.is_gestor() or request.user.is_administrador() or request.user.is_analista()):
        return redirect('dashboard')
    dept = request.session.get('current_department_obj') or request.user.department
    return render(request, 'core/auditoria_atendimentos.html', {
        'title': 'Auditoria de Atendimentos', 'department': dept, 'is_admin': request.user.is_administrador(),
        'is_gestor': request.user.is_gestor(), 'is_analista': request.user.is_analista(),
    })

@login_required
def api_get_system_notifications(request):
    notifications = SystemNotification.objects.filter(is_active=True).order_by('-created_at')[:5]
    return JsonResponse({'notifications': [{
        'id': n.id, 'title': n.title, 'message': n.message, 'details': n.details,
        'category': n.get_category_display(), 'category_code': n.category,
        'created_at': n.created_at.strftime('%d/%m/%Y %H:%M'), 'timestamp': n.created_at.timestamp()
    } for n in notifications]})
